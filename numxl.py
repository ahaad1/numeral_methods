import os 
from array import array
from openpyxl import Workbook
import pandas as pd
import numpy as np
import matplotlib as mp

def create_xlsx(filename: str, cols: int):
    try:
        wb = Workbook()
        xlsx_sheet = wb.active
        for col in range(1, cols + 100):
            xlsx_sheet.cell(row=1, column=col, value=col)
        file_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), f"{filename}.xlsx")
        wb.save(file_path)
        print(f"Файл создан и сохранен: {file_path}")
    
    except Exception as e:
        print(f"Ошибка создания файла: {e}")



def write_xlsx(arr: list, row: int, col: int, sheet, message: str = None):
    """
    Записывает массив arr на лист sheet, начиная с позиции (row, col).
    Если передано сообщение message, оно будет записано перед массивом.
    
    :param arr: Двумерный массив (список списков), который нужно записать.
    :param row: Начальная строка для записи.
    :param col: Начальный столбец для записи.
    :param sheet: Лист Excel-документа (объект openpyxl).
    :param message: Сообщение, которое нужно записать перед массивом (опционально).
    """
    try:
        if message:  # Если сообщение передано
            sheet.cell(row=row, column=col, value=message)
            row += 1  # Сместить начальную строку для массива вниз
        
        for i, line in enumerate(arr):
            for j, value in enumerate(line):
                sheet.cell(row=row + i, column=col + j, value=value)

    except Exception as e:
        print(f"Ошибка записи массива: {e}")