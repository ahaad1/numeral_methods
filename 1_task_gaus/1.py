import random
import os 
import openpyxl
from array import array
from openpyxl import Workbook
import pandas as pd
import numpy as np
import matplotlib as mp
# from pprint import pprint

def init_xlsx(filename: str, cols: int):
    try:
        wb = Workbook()
        xlsx_sheet = wb.active
        for col in range(1, cols + 100):
            xlsx_sheet.cell(row=1, column=col, value=col)
        file_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), f"{filename}.xlsx")
        wb.save(file_path)
        print(f"Файл создан и сохранен: {file_path}")
    
    except Exception as e:
        print(f"Ошибка создания файла: {e}")

def generate_matrix(xlsx_filename:str, rows:int, cols:int, k_student_num:int):
    try:
        book = openpyxl.open(f"{os.path.dirname(os.path.realpath(__file__))}/{xlsx_filename}.xlsx", read_only=False)
        xlsx_sheet = book.active
    except:
        print(f"ошибка чтения файла {os.path.dirname(os.path.realpath(__file__))}/{xlsx_filename}.xlsx")

    try:
        for row in range(0, rows):
            xlsx_sheet[row + 2][rows].value = (20 - k_student_num)/(row + 2)
            for col in range(0, cols):
                xlsx_sheet[row + 2][col].value = (k_student_num + 1)/(row + col + 22 - k_student_num)
        book.save(f"{os.path.dirname(os.path.realpath(__file__))}/{xlsx_filename}.xlsx")
        print(f"Матрица {rows}x{cols} сгенерирована и сохранена в {os.path.dirname(os.path.realpath(__file__))}/{xlsx_filename}.xlsx")
    except Exception as e:
        print(f"ошибка записи в файл {os.path.dirname(os.path.realpath(__file__))}/{xlsx_filename}.xlsx")
        print(e)

def read_matrix(xlsx_filename:str, rows:int, cols:int) -> list:
    try:
        book = openpyxl.load_workbook(f"{os.path.dirname(os.path.realpath(__file__))}/{xlsx_filename}.xlsx")
        book_sheet = book.active
    except Exception as e:
        print(f"ошибка чтения файла {os.path.dirname(os.path.realpath(__file__))}/{xlsx_filename}.xlsx")
        print(f"{e}")

    matrix = []

    for row in range(2, rows + 2):
        row_list = []
        for col in range(2, cols + 2):
            cell = book_sheet.cell(row=row, column=col)
            row_list.append(cell.value)
        matrix.append(row_list)
    print(f"матрица успешно считана")

    return matrix

def forward_elimination(matrix: list, rows: int, cols: int):
    for i in range(rows - 1): 
        for j in range(i + 1, rows): 
            coef = matrix[j][i] / matrix[i][i]
            for k in range(i, cols):
                matrix[j][k] -= coef * matrix[i][k]
            print(f"Row {j}, Column {i}: Coef = {coef}")
        print(f"\n\nMatrix after step {i + 1}:", *[str(row) for row in matrix], sep="\n")
    return matrix

# def gauss_method_solution(xlsx_filename:str, matrix:list, row:int, cols:int) -> list:
    



def main():
    init_xlsx('gaus', 6)
    generate_matrix('gaus', 6, 6, 2)
    matrix = read_matrix('gaus', 6, 6)

    print("[")
    print(*[str(row) for row in matrix], sep=",\n")
    print("]")

    matrix = forward_elimination(matrix=matrix, rows=6, cols=6)
main()